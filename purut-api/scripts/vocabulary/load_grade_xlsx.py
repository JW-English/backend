#!/usr/bin/env python3
"""학년별 단어 xlsx → 적재 SQL.

사용:
    python3 load_grade_xlsx.py \
        --grade1 ~/Downloads/고1....xlsx \
        --grade2 ~/Downloads/고2....xlsx \
        --grade3 ~/Downloads/고3....xlsx > out/grades.sql

시트 '전체 어휘' 의 A~H 열만 쓴다. I 열부터는 출제 통계라 학습에 필요 없다.

    A 전체 순번   B Day   C Day 내 순번   D 표제어
    E 대표 품사   F 한글 뜻   G 영문 예문   H 예문 해석

## 같은 단어가 학년마다 나온다

고2·고3 에 1450개가 겹치는데 뜻·예문은 대개 다르다. words 는
(headword, meaning_ko) 가 유니크라, 뜻이 다르면 행이 갈리고 같으면 한 행을
공유한다. 학년별 배치는 word_day_items 가 들고 있으므로 이 구조로 문제가 없다 —
고2 학생과 고3 학생이 같은 단어를 각자의 뜻으로 본다.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

SHEET = "전체 어휘"
DAY = re.compile(r"Day\s*0*(\d+)", re.IGNORECASE)


def quote(value) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).strip().replace("'", "''") + "'"


def read(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"'{SHEET}' 시트가 없습니다: {path}")

    rows = []
    for r in wb[SHEET].iter_rows(min_row=2, values_only=True):
        if not r or not r[3]:
            continue
        m = DAY.search(str(r[1] or ""))
        if not m:
            continue
        rows.append({
            "day": int(m.group(1)),
            "order": int(r[2] or 0),
            "headword": str(r[3]).strip(),
            "pos": (str(r[4]).strip() if r[4] else None),
            "meaning": str(r[5]).strip(),
            "example_en": (str(r[6]).strip() if r[6] else None),
            "example_ko": (str(r[7]).strip() if r[7] else None),
        })
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--grade1", required=True)
    parser.add_argument("--grade2", required=True)
    parser.add_argument("--grade3", required=True)
    args = parser.parse_args()

    sources = [("GRADE_1", args.grade1), ("GRADE_2", args.grade2), ("GRADE_3", args.grade3)]

    print("-- 생성물입니다. load_grade_xlsx.py 로 다시 만들 수 있습니다.")
    print("BEGIN;")

    # 뜻이 같은 단어는 한 행을 공유한다. 중복 INSERT 를 미리 접어둔다
    seen_words: set[tuple[str, str]] = set()
    summary = []

    for level, path in sources:
        rows = read(path)
        days = sorted({r["day"] for r in rows})
        summary.append((level, Path(path).name, len(rows), len(days)))

        print(f"\n-- ══ {level} · {Path(path).name} ══")
        for day in days:
            print(
                "INSERT INTO word_days (level, day_no, title, scheduled_date) "
                f"VALUES ({quote(level)}, {day}, {quote(f'DAY {day:02d}')}, current_date) "
                "ON CONFLICT (level, day_no) DO UPDATE SET "
                "title = EXCLUDED.title, scheduled_date = EXCLUDED.scheduled_date;"
            )

        for r in rows:
            key = (r["headword"], r["meaning"])
            if key not in seen_words:
                seen_words.add(key)
                print(
                    "INSERT INTO words (headword, meaning_ko, part_of_speech, example_en, example_ko) "
                    f"VALUES ({quote(r['headword'])}, {quote(r['meaning'])}, {quote(r['pos'])}, "
                    f"{quote(r['example_en'])}, {quote(r['example_ko'])}) "
                    "ON CONFLICT (headword, meaning_ko) DO UPDATE SET "
                    "part_of_speech = EXCLUDED.part_of_speech, example_en = EXCLUDED.example_en, "
                    "example_ko = EXCLUDED.example_ko;"
                )

            # 뜻까지 지정해 정확한 행에 건다. headword 만으로 찾으면 학년마다 뜻이
            # 다른 단어에서 엉뚱한 행이 걸린다
            print(
                "INSERT INTO word_day_items (day_id, word_id, sort_order) "
                f"SELECT d.id, w.id, {r['order']} FROM word_days d, words w "
                f"WHERE d.level = {quote(level)} AND d.day_no = {r['day']} "
                f"AND w.headword = {quote(r['headword'])} AND w.meaning_ko = {quote(r['meaning'])} "
                "ON CONFLICT (day_id, word_id) DO UPDATE SET sort_order = EXCLUDED.sort_order;"
            )

    print("\nCOMMIT;")

    print("\n── 적재 요약 ──", file=sys.stderr)
    for level, name, n, days in summary:
        print(f"  {level}  {n}단어 · DAY {days}  ({name})", file=sys.stderr)
    print(f"  words 행 {len(seen_words)}개 (뜻이 같은 단어는 공유)", file=sys.stderr)


if __name__ == "__main__":
    main()
