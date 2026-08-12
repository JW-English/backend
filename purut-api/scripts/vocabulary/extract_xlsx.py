#!/usr/bin/env python3
"""어휘 xlsx → 중간 CSV.

사용:
    python3 extract_xlsx.py --input intermediate.xlsx --set intermediate > data/intermediate.csv

기대하는 시트 구조 (열 4개):
    단원 | 유형 | 영단어 | 한글 뜻

시트가 둘이면 두 번째(축약본)를 meaning_ko, 첫 번째(전체)를 meanings 로 쓴다.
축약본이 퀴즈 선택지 길이에 맞고, 전체본은 단어장 화면에 쓴다.
시트가 하나면 같은 값을 둘 다에 넣는다.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

DAY_RE = re.compile(r"DAY\s*0*(\d+)", re.IGNORECASE)
# 표제어 / 파생어. 그대로 tags 에 실어 보내 퀴즈에서 파생어를 뺄지 고를 수 있게 한다
TYPE_MAP = {"표": "headword", "파": "derived"}


def read_sheet(ws) -> list[dict]:
    rows = []
    for unit, wtype, word, meaning in ws.iter_rows(min_row=2, values_only=True):
        if not word:
            continue
        rows.append(
            {
                "unit": (unit or "").strip(),
                "type": TYPE_MAP.get((wtype or "").strip(), ""),
                "headword": str(word).strip(),
                "meaning": (meaning or "").strip(),
            }
        )
    return rows


def day_of(unit: str, extra: dict[str, int]) -> tuple[int, str]:
    """'DAY 01' → (1, '')  ·  특수 단원 → (61.., 원래 이름)

    day_no 가 정수 컬럼이라 특수 단원에도 번호를 줘야 한다.
    번호는 등장 순서대로 뒤에 이어 붙이고 이름은 title 에 남긴다.
    """
    m = DAY_RE.search(unit)
    if m:
        return int(m.group(1)), ""

    if unit not in extra:
        extra[unit] = max(extra.values(), default=0) + 1
    return extra[unit], unit


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--set", required=True, help="beginner | intermediate | advanced")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    sheets = wb.sheetnames

    full = read_sheet(wb[sheets[0]])
    short = read_sheet(wb[sheets[1]]) if len(sheets) > 1 else full

    if len(full) != len(short):
        print(
            f"⚠️  시트 행 수가 다릅니다 ({len(full)} vs {len(short)}). "
            "첫 시트 기준으로 맞추고 짝이 없으면 전체본을 씁니다.",
            file=sys.stderr,
        )

    # 축약본을 표제어로 찾아 쓴다. 행 순서가 어긋나도 맞도록
    short_by_word = {r["headword"]: r["meaning"] for r in short}

    # 특수 단원 번호는 일반 DAY 최대값 다음부터
    max_day = 0
    for r in full:
        m = DAY_RE.search(r["unit"])
        if m:
            max_day = max(max_day, int(m.group(1)))
    extra: dict[str, int] = {}
    base = max_day

    writer = csv.writer(sys.stdout)
    writer.writerow(
        ["set", "day_no", "day_title", "sort_order", "headword", "word_type",
         "meaning_ko", "meanings"]
    )

    order: dict[int, int] = {}
    for r in full:
        day_no, title = day_of(r["unit"], extra)
        if title:
            day_no += base

        order[day_no] = order.get(day_no, 0) + 1

        meanings = r["meaning"]
        meaning_ko = short_by_word.get(r["headword"], meanings)

        writer.writerow(
            [args.set, day_no, title, order[day_no], r["headword"], r["type"],
             meaning_ko, meanings]
        )

    print(
        f"✅ {len(full)}단어 · DAY {len(order)}개 (특수 단원 {len(extra)}개)",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
